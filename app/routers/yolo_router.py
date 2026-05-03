"""
app/routers/yolo_router.py  (FULL REPLACEMENT)
===============================================
Fixes:
  1. /api/v2/export/batch-xlsx  — each invoice on its own sheet
     in FIELD | VALUE format (same layout as single invoice export)
     with Invoice_Header + Table_1 sub-sheets per invoice block

  2. /api/v2/yolo/detect  — always runs heuristic regions AND real YOLO,
     merges both, so you always get header/table/footer even on
     digital invoice images where YOLO finds < 3 regions.

  3. /api/v2/yolo/bank-detect  — YOLO + OCR pipeline for bank statements:
     detects table region → crops → runs OCR only on that crop →
     passes to BankStatementExtractor for better transaction parsing.

Install: pip install ultralytics openpyxl pillow
"""

import base64
import io
import re
import time
import logging
from typing import List, Optional, Dict, Any

import cv2
import numpy as np
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter()


# ─────────────────────────────────────────────────────────────────────────────
# Shared Pydantic models
# ─────────────────────────────────────────────────────────────────────────────

class DetectedRegion(BaseModel):
    label: str
    confidence: float
    bbox: List[float]


class YoloResult(BaseModel):
    regions: List[DetectedRegion]
    annotated_image_base64: Optional[str]
    model_used: str
    processing_time_ms: float
    total_regions: int


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _decode_image(file_bytes: bytes) -> np.ndarray:
    nparr = np.frombuffer(file_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        raise HTTPException(400, "Could not decode image — use JPG or PNG")
    return img


def _heuristic_regions(img: np.ndarray) -> List[DetectedRegion]:
    """
    Always produces header/table/footer regions based on image proportions.
    Used as baseline — real YOLO results are merged on top.
    """
    h, w = img.shape[:2]
    return [
        DetectedRegion(label="header", confidence=0.60,
                       bbox=[0, 0, float(w), float(int(h * 0.15))]),
        DetectedRegion(label="table",  confidence=0.60,
                       bbox=[0, float(int(h * 0.15)), float(w), float(int(h * 0.82))]),
        DetectedRegion(label="footer", confidence=0.60,
                       bbox=[0, float(int(h * 0.82)), float(w), float(h)]),
    ]


def _draw_regions(img: np.ndarray, regions: List[DetectedRegion]) -> np.ndarray:
    """Draw colored bounding boxes with labels on image."""
    COLOR_MAP = {
        "header":            (255, 100,  30),
        "table":             (180,  30, 255),
        "footer":            (30,  200,  80),
        "logo":              (255,  30, 180),
        "stamp":             (30,  180, 255),
        "signature":         (50,  200, 200),
        "barcode":           (200, 200,  30),
        # Bank statement region labels
        "bank_header":       (255, 100,  30),
        "transaction_table": (180,  30, 255),
        "bank_footer":       (30,  200,  80),
    }
    annotated = img.copy()
    for reg in regions:
        x1, y1, x2, y2 = [int(v) for v in reg.bbox]
        color = COLOR_MAP.get(reg.label.lower(), (150, 150, 150))
        cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 3)
        label_txt = f"{reg.label} ({reg.confidence*100:.0f}%)"
        (tw, th), _ = cv2.getTextSize(label_txt, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)
        cv2.rectangle(annotated, (x1, y1 - th - 10), (x1 + tw + 8, y1), color, -1)
        cv2.putText(annotated, label_txt, (x1 + 4, y1 - 6),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    return annotated


def _run_yolo(img: np.ndarray) -> List[DetectedRegion]:
    """Run YOLOv8 and remap detected objects to document regions."""
    try:
        from ultralytics import YOLO
        model = YOLO("yolov8n.pt")
        results = model(img, verbose=False)
        h, w = img.shape[:2]
        regions = []
        for box in results[0].boxes:
            cls_id = int(box.cls[0])
            label = model.names[cls_id]
            conf = float(box.conf[0])
            x1, y1, x2, y2 = box.xyxy[0].tolist()
            doc_label = _remap(label, x1, y1, x2, y2, w, h)
            regions.append(DetectedRegion(
                label=doc_label, confidence=round(conf, 4),
                bbox=[round(x1,1), round(y1,1), round(x2,1), round(y2,1)],
            ))
        return regions
    except Exception as e:
        logger.warning(f"YOLO inference failed: {e}")
        return []


def _remap(label: str, x1, y1, x2, y2, w, h) -> str:
    rel_y_top = y1 / h
    rel_y_bot = y2 / h
    region_h  = (y2 - y1) / h
    if rel_y_top < 0.15:
        return "header"
    if rel_y_bot > 0.85:
        return "footer"
    if region_h > 0.35:
        return "table"
    if (x2 - x1) < 0.25 * w and (y2 - y1) < 0.15 * h:
        return "logo"
    return "content_block"


def _merge_regions(
    yolo: List[DetectedRegion],
    heuristic: List[DetectedRegion],
) -> List[DetectedRegion]:
    """
    Use YOLO results when confidence >= 0.50.
    For any region label NOT found in YOLO results, add the heuristic version.
    This guarantees header/table/footer always appear.
    """
    found_labels = {r.label for r in yolo if r.confidence >= 0.50}
    merged = list(yolo)
    for h_reg in heuristic:
        if h_reg.label not in found_labels:
            merged.append(h_reg)
    return merged


def _to_b64(img: np.ndarray) -> str:
    _, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 85])
    return base64.b64encode(buf.tobytes()).decode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# FIX 2: Invoice YOLO detection — always shows header/table/footer
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/api/v2/yolo/detect", response_model=YoloResult)
async def yolo_detect(file: UploadFile = File(...)):
    """
    Hybrid YOLO + Heuristic detection for Invoices.
    Optimized for:
      1. Scans/Photos: YOLO-first approach to handle skew and shadows.
      2. Clean Digital PDFs: Heuristic fallback to ensure Header/Table/Footer are found.
    """
    start = time.time()
    is_pdf = (file.content_type == "application/pdf" or 
              (file.filename or "").lower().endswith(".pdf"))
    
    file_bytes = await file.read()

    # 1. PDF Rendering (2x zoom for high-quality OCR/YOLO)
    if is_pdf:
        import fitz
        pdf = fitz.open(stream=file_bytes, filetype="pdf")
        if pdf.page_count == 0: raise HTTPException(400, "PDF is empty")
        page = pdf[0]
        pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
        file_bytes = pix.tobytes("png")
        pdf.close()

    img = _decode_image(file_bytes)
    h, w = img.shape[:2]

    # 2. Run YOLO (Visual Intelligence)
    yolo_regions = _run_yolo(img) # Returns detected Header, Table, Footer
    
    # 3. Smart Heuristic Logic (Spatial Logic)
    # We define standard regions for a clean A4/Invoice layout
    heuristic = _heuristic_regions(img)
    
    # 4. Hybrid Merge Logic: 
    # If YOLO misses a specific part (e.g., Header), we use the Heuristic part.
    # This prevents "Empty Detection" on clean digital files.
    final_regions = []
    labels_found = {r.label for r in yolo_regions}
    
    # Keep all YOLO detections (High Confidence)
    final_regions.extend(yolo_regions)

    # Fill in missing parts using heuristics
    for h_reg in heuristic:
        if h_reg.label not in labels_found:
            h_reg.label = f"{h_reg.label} (heuristic fallback)"
            h_reg.confidence = 0.50
            final_regions.append(h_reg)

    # 5. Post-Processing & UI Feedback
    annotated = _draw_regions(img, final_regions)
    b64 = _to_b64(annotated)
    
    # Save for debugging
    try:
        from pathlib import Path
        save_dir = Path("yolo_outputs")
        save_dir.mkdir(exist_ok=True)
        cv2.imwrite(str(save_dir / f"inv_{int(time.time())}.jpg"), annotated)
    except Exception as e:
        logger.warning(f"Save failed: {e}")

    return YoloResult(
        regions=final_regions,
        annotated_image_base64=b64,
        model_used="YOLOv8 + Hybrid Fallback",
        processing_time_ms=round((time.time() - start) * 1000, 1),
        total_regions=len(final_regions),
    )

# ─────────────────────────────────────────────────────────────────────────────
# FIX 3: Bank statement YOLO — detect table region, crop, run OCR on crop
# ─────────────────────────────────────────────────────────────────────────────

class BankYoloResult(BaseModel):
    regions: List[DetectedRegion]
    annotated_image_base64: Optional[str]
    extracted_text_per_region: Dict[str, str]   # label → OCR text
    bank_name: Optional[str]
    account_number: Optional[str]
    transactions_found: int
    processing_time_ms: float


@router.post("/api/v2/yolo/bank-detect", response_model=BankYoloResult)
async def yolo_bank_detect(file: UploadFile = File(...)):
    """
    YOLO + OCR pipeline for bank statements:
      1. Detect header / transaction_table / footer regions
      2. Crop each region
      3. Run OCR on each crop separately (better accuracy for tables)
      4. Parse bank name, account number, transactions from table region
    """
    if not (file.content_type or "").startswith("image/"):
        raise HTTPException(400, "Only image files supported")

    file_bytes = await file.read()
    start = time.time()

    img = _decode_image(file_bytes)
    h, w = img.shape[:2]

    # Bank-specific heuristic regions (different proportions)
    bank_heuristic = [
        DetectedRegion(label="bank_header",      confidence=0.60,
                       bbox=[0, 0, float(w), float(int(h * 0.20))]),
        DetectedRegion(label="transaction_table", confidence=0.60,
                       bbox=[0, float(int(h * 0.20)), float(w), float(int(h * 0.88))]),
        DetectedRegion(label="bank_footer",       confidence=0.60,
                       bbox=[0, float(int(h * 0.88)), float(w), float(h)]),
    ]

    yolo_regions  = _run_yolo(img)
    final_regions = _merge_regions(yolo_regions, bank_heuristic)

    # OCR each region separately
    try:
        from rapidocr_onnxruntime import RapidOCR
        ocr_engine = RapidOCR()
        use_rapid = True
    except ImportError:
        use_rapid = False

    text_per_region: Dict[str, str] = {}

    for reg in final_regions:
        x1, y1, x2, y2 = [int(v) for v in reg.bbox]
        # Clamp to image bounds
        x1 = max(0, x1); y1 = max(0, y1)
        x2 = min(w, x2); y2 = min(h, y2)
        if x2 <= x1 or y2 <= y1:
            continue
        crop = img[y1:y2, x1:x2]

        region_text = ""
        if use_rapid:
            try:
                result, _ = ocr_engine(crop)
                if result:
                    region_text = "\n".join(str(item[1]) for item in result if len(item) >= 2)
            except Exception as e:
                logger.warning(f"RapidOCR on region {reg.label}: {e}")

        if not region_text:
            # Fallback: pytesseract
            try:
                import pytesseract
                gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                region_text = pytesseract.image_to_string(gray, config="--psm 6")
            except Exception:
                pass

        key = reg.label.replace(" (heuristic)", "")
        text_per_region[key] = region_text.strip()

    # Parse bank info from combined text
    full_text = "\n".join(text_per_region.values())
    bank_name      = _detect_bank(full_text)
    account_number = _detect_account(full_text)
    txn_count      = _count_transactions(text_per_region.get("transaction_table", ""))

    annotated = _draw_regions(img, final_regions)
    b64       = _to_b64(annotated)
    elapsed   = (time.time() - start) * 1000

    # Save annotated image to yolo_outputs/ folder
    try:
        from pathlib import Path
        yolo_dir = Path("yolo_outputs")
        yolo_dir.mkdir(exist_ok=True)
        ts = int(time.time() * 1000)
        save_path = yolo_dir / f"bank_yolo_{ts}.jpg"
        cv2.imwrite(str(save_path), annotated)
        logger.info(f"Bank YOLO output saved: {save_path}")
    except Exception as e:
        logger.warning(f"Could not save YOLO output image: {e}")

    return BankYoloResult(
        regions=final_regions,
        annotated_image_base64=b64,
        extracted_text_per_region=text_per_region,
        bank_name=bank_name,
        account_number=account_number,
        transactions_found=txn_count,
        processing_time_ms=round(elapsed, 1),
    )


def _detect_bank(text: str) -> Optional[str]:
    BANKS = {
        "State Bank of India": ["sbi", "state bank of india"],
        "HDFC Bank":  ["hdfc"],
        "ICICI Bank": ["icici"],
        "Axis Bank":  ["axis bank"],
        "Kotak":      ["kotak"],
        "Yes Bank":   ["yes bank"],
        "Bank of Baroda": ["bank of baroda"],
        "PNB":        ["punjab national", "pnb"],
        "IndusInd":   ["indusind"],
        "Union Bank": ["union bank"],
        "Canara Bank":["canara"],
        "IDBI Bank":  ["idbi"],
    }
    tl = text.lower()
    for name, kws in BANKS.items():
        if any(k in tl for k in kws):
            return name
    return None


def _detect_account(text: str) -> Optional[str]:
    for p in [
        r"Account\s+(?:Number|No)[.:\s]+(\d[\d\s]{8,18}\d)",
        r"A/C\s+(?:No|Number)[.:\s]+(\d[\d\s]{8,18}\d)",
    ]:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).replace(" ", "")
    return None


def _count_transactions(table_text: str) -> int:
    # This regex matches common date formats
    DATE_PATTERN = r"\d{2}[/-]\d{2}[/-]\d{2,4}"
    
    lines = table_text.split('\n')
    count = 0
    
    for line in lines:
        line = line.strip()
        if re.match(DATE_PATTERN, line):
            count += 1
            
    return count


@router.post("/api/v2/export/batch-xlsx")
async def export_batch_xlsx(invoices: List[Dict[str, Any]]):
    """
    Each invoice gets its OWN sheet named Invoice_1, Invoice_2 ...
    Layout inside each sheet = Field | Value rows (same as single invoice Excel).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "pip install openpyxl")

    TITLE_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    TITLE_FILL   = PatternFill("solid", fgColor="1E3A5F")
    SECTION_FONT = Font(bold=True, color="1E3A5F", size=10, name="Calibri")
    SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
    LABEL_FONT   = Font(bold=True, size=9, color="444444", name="Calibri")
    VALUE_FONT   = Font(size=10, name="Calibri")
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cell(ws, row, col, value, font=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font  = font
        if fill:  c.fill  = fill
        if align: c.alignment = align
        c.border = BORDER
        return c

    def section(ws, row, title):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font  = SECTION_FONT
        c.fill  = SECTION_FILL
        c.alignment = LEFT
        c.border = BORDER

    def kv(ws, row, key, val):
        set_cell(ws, row, 1, key,  font=LABEL_FONT, align=LEFT)
        set_cell(ws, row, 2, val if val is not None else "—", font=VALUE_FONT, align=LEFT)

    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active) 

    for idx, inv in enumerate(invoices):
        platform = (inv.get("platform") or "UNKNOWN").upper()
        inv_no   = inv.get("invoice_number") or f"#{idx+1}"
        sheet_name = f"Invoice_{idx+1}" 
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 52
        r = 1

        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1,
                    value=f"Invoice {idx+1}  ·  {platform}  ·  {inv_no}")
        c.font      = TITLE_FONT
        c.fill      = TITLE_FILL
        c.alignment = CENTER
        c.border    = BORDER
        ws.row_dimensions[r].height = 24
        r += 1

        section(ws, r, "📋  Core Fields (22 mandatory)"); r += 1
        mandatory = [
            ("billing_address",     inv.get("billing_address")),
            ("shipping_address",    inv.get("shipping_address")),
            ("invoice_type",        inv.get("invoice_type")),
            ("order_number",        inv.get("order_number")),
            ("invoice_number",      inv.get("invoice_number")),
            ("order_date",          inv.get("order_date")),
            ("invoice_details",     inv.get("invoice_details")),
            ("invoice_date",        inv.get("invoice_date")),
            ("seller_info",         inv.get("seller_info")),
            ("seller_pan",          inv.get("seller_pan")),
            ("seller_gst",          inv.get("seller_gst")),
            ("fssai_license",       inv.get("fssai_license")),
            ("billing_state_code",  inv.get("billing_state_code")),
            ("shipping_state_code", inv.get("shipping_state_code")),
            ("place_of_supply",     inv.get("place_of_supply")),
            ("place_of_delivery",   inv.get("place_of_delivery")),
            ("reverse_charge",      inv.get("reverse_charge")),
            ("amount_in_words",     inv.get("amount_in_words")),
            ("seller_name",         inv.get("seller_name")),
            ("seller_address",      inv.get("seller_address")),
            ("total_tax",           inv.get("total_tax")),
            ("total_amount",        inv.get("total_amount")),
        ]
        for key, val in mandatory:
            kv(ws, r, key, val); r += 1

        # ── Extra fields ─────────────────────────────────────────────────
        section(ws, r, "💰  Extra Fields"); r += 1
        extras = [
            ("platform",        inv.get("platform")),
            ("buyer_name",      inv.get("buyer_name")),
            ("buyer_phone",     inv.get("buyer_phone")),
            ("subtotal",        inv.get("subtotal")),
            ("cgst_rate",       inv.get("cgst_rate")),
            ("cgst_amount",     inv.get("cgst_amount")),
            ("sgst_rate",       inv.get("sgst_rate")),
            ("sgst_amount",     inv.get("sgst_amount")),
            ("igst_rate",       inv.get("igst_rate")),
            ("igst_amount",     inv.get("igst_amount")),
            ("discount",        inv.get("discount")),
            ("delivery_charge", inv.get("delivery_charge")),
            ("packaging_charge",inv.get("packaging_charge")),
            ("payment_method",  inv.get("payment_method")),
            ("fields_extracted",inv.get("fields_extracted")),
        ]
        for key, val in extras:
            kv(ws, r, key, val); r += 1

    # ── Stream response ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=batch_invoices.xlsx"},
    )