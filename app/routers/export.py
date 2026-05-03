from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from app.models.schemas import InvoiceData, ExtractionResponse
from typing import List, Optional
import csv
import io
import json
import os
import tempfile
import logging
from pathlib import Path

EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

logger = logging.getLogger(__name__)
router = APIRouter()



MENTOR_FIELDS = [
    ("billing_address", "billing_address"),
    ("shipping_address", "shipping_address"),
    ("invoice_type", "invoice_type"),
    ("order_number", "order_number"),
    ("invoice_number", "invoice_number"),
    ("order_date", "order_date"),
    ("invoice_details", "invoice_details"),
    ("invoice_date", "invoice_date"),
    ("seller_info", "seller_info"),
    ("seller_pan", "seller_pan"),
    ("seller_gst", "seller_gst"),
    ("fssai_license", "fssai_license"),
    ("billing_state_code", "billing_state_code"),
    ("shipping_state_code", "shipping_state_code"),
    ("place_of_supply", "place_of_supply"),
    ("place_of_delivery", "place_of_delivery"),
    ("reverse_charge", "reverse_charge"),
    ("amount_in_words", "amount_in_words"),
    ("seller_name", "seller_name"),
    ("seller_address", "seller_address"),
    ("total_tax", "total_tax"),
    ("total_amount", "total_amount"),
]

EXTRA_FIELDS = [
    ("platform", "platform"),
    ("buyer_name", "buyer_name"),
    ("buyer_phone", "buyer_phone"),
    ("subtotal", "subtotal"),
    ("cgst_rate", "cgst_rate"),
    ("cgst_amount", "cgst_amount"),
    ("sgst_rate", "sgst_rate"),
    ("sgst_amount", "sgst_amount"),
    ("igst_rate", "igst_rate"),
    ("igst_amount", "igst_amount"),
    ("discount", "discount"),
    ("delivery_charge", "delivery_charge"),
    ("packaging_charge", "packaging_charge"),
    ("payment_method", "payment_method"),
]


def _get_value(data: InvoiceData, field_name: str) -> str:
    """Get a field value from InvoiceData, converting to string."""
    val = getattr(data, field_name, None)
    if val is None:
        return ""
    if hasattr(val, "value"):  # Enum
        return str(val.value)
    return str(val)


@router.post("/xlsx")
async def export_xlsx(invoice: InvoiceData):

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            500,
            "openpyxl not installed. Run: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice_Header"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws["A1"] = "Field"
    ws["B1"] = "Value"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = 2
    for field_label, field_name in MENTOR_FIELDS:
        ws[f"A{row}"] = field_label
        ws[f"B{row}"] = _get_value(invoice, field_name)
        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    row += 1

    ws[f"A{row}"] = "Extra Fields"
    ws[f"B{row}"] = ""
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="4472C4")
    row += 1

    for field_label, field_name in EXTRA_FIELDS:
        ws[f"A{row}"] = field_label
        ws[f"B{row}"] = _get_value(invoice, field_name)
        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 2: Table_1 — Line Items (matches professor's Output Template)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws2 = wb.create_sheet(title="Table_1")

    table_headers = [
        ("A", "Sl.", 6),
        ("B", "Description", 45),
        ("C", "Unit Price", 12),
        ("D", "Discount", 12),
        ("E", "Qty", 6),
        ("F", "Net Amount", 14),
        ("G", "Tax Rate", 10),
        ("H", "Tax Type", 10),
        ("I", "Tax Amount", 12),
        ("J", "Total", 14),
    ]

    for col_letter, header_text, width in table_headers:
        ws2.column_dimensions[col_letter].width = width
        cell = ws2[f"{col_letter}1"]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Fill line items
    items = invoice.items or []
    item_row = 2
    for idx, item in enumerate(items, 1):
        ws2[f"A{item_row}"] = idx
        ws2[f"B{item_row}"] = getattr(item, "description", "") or ""
        ws2[f"C{item_row}"] = getattr(item, "unit_price", "") or ""
        ws2[f"D{item_row}"] = getattr(item, "discount", "") or ""
        ws2[f"E{item_row}"] = getattr(item, "quantity", "") or ""
        ws2[f"F{item_row}"] = getattr(item, "total_price", "") or ""
        ws2[f"G{item_row}"] = getattr(item, "tax_rate", "") or ""
        ws2[f"H{item_row}"] = ""
        ws2[f"I{item_row}"] = ""
        ws2[f"J{item_row}"] = getattr(item, "total_price", "") or ""
        for col in "ABCDEFGHIJ":
            ws2[f"{col}{item_row}"].border = thin_border
        item_row += 1

    # Ensure at least 3 empty item rows (professor's template has 3 placeholder rows)
    while item_row < 5:
        ws2[f"A{item_row}"] = item_row - 1
        for col in "ABCDEFGHIJ":
            ws2[f"{col}{item_row}"].border = thin_border
        item_row += 1

    # TOTAL row
    ws2[f"A{item_row}"] = "TOTAL:"
    ws2[f"A{item_row}"].font = Font(bold=True)
    ws2[f"J{item_row}"] = _get_value(invoice, "total_amount")
    ws2[f"J{item_row}"].font = Font(bold=True)
    for col in "ABCDEFGHIJ":
        ws2[f"{col}{item_row}"].border = thin_border

    inv_num = invoice.invoice_number or "extracted"
    platform = invoice.platform.value if hasattr(invoice.platform, "value") else str(invoice.platform)
    filename = f"{platform}_{inv_num}.xlsx"
    file_path = EXPORTS_DIR / filename
    wb.save(file_path)

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/xlsx-batch")
async def export_xlsx_batch(invoices: List[InvoiceData]):

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"

    all_fields = ["platform"] + [f[0] for f in MENTOR_FIELDS] + [f[0] for f in EXTRA_FIELDS[1:]]
    for col, header in enumerate(all_fields, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for row_idx, inv in enumerate(invoices, 2):
        platform_val = inv.platform.value if hasattr(inv.platform, "value") else str(inv.platform)
        ws_summary.cell(row=row_idx, column=1, value=platform_val)
        for col, (_, field_name) in enumerate(MENTOR_FIELDS, 2):
            ws_summary.cell(row=row_idx, column=col, value=_get_value(inv, field_name))
        offset = len(MENTOR_FIELDS) + 2
        for col, (_, field_name) in enumerate(EXTRA_FIELDS[1:], offset):
            ws_summary.cell(row=row_idx, column=col, value=_get_value(inv, field_name))

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, inv in enumerate(invoices):
        platform_val = inv.platform.value if hasattr(inv.platform, "value") else str(inv.platform)
        inv_num = inv.invoice_number or f"invoice_{i+1}"
        sheet_name = f"{platform_val}_{inv_num}"[:31]  

        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60

        ws["A1"] = "Field"
        ws["B1"] = "Value"
        for cell in [ws["A1"], ws["B1"]]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            cell.border = thin_border

        ws["A2"] = "platform"
        ws["B2"] = platform_val
        ws["A2"].border = thin_border
        ws["B2"].border = thin_border

        row = 3
        for field_label, field_name in MENTOR_FIELDS:
            ws[f"A{row}"] = field_label
            ws[f"B{row}"] = _get_value(inv, field_name)
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1

    file_path = EXPORTS_DIR / "batch_extraction_results.xlsx"
    wb.save(file_path)

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="batch_extraction_results.xlsx",
    )



@router.post("/csv")
async def export_csv(invoices: List[InvoiceData]):
    """Export multiple invoices as a CSV file (one row per invoice)."""
    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["platform"] + [f[0] for f in MENTOR_FIELDS] + [f[0] for f in EXTRA_FIELDS[1:]]
    writer.writerow(headers)

    for inv in invoices:
        platform_val = inv.platform.value if hasattr(inv.platform, "value") else str(inv.platform)
        row = [platform_val]
        for _, field_name in MENTOR_FIELDS:
            row.append(_get_value(inv, field_name))
        for _, field_name in EXTRA_FIELDS[1:]:
            row.append(_get_value(inv, field_name))
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=invoices_export.csv"},
    )