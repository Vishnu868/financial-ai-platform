"""
Universal Bank Statement Extraction Endpoint — v2.0
====================================================
Handles:
  • Single image  (JPG / PNG / TIFF / BMP / WEBP)
  • Single-page PDF
  • Multi-page PDF  ← NEW: each page extracted independently, merged
  • Any bank layout (HDFC, SBI, ICICI, Axis, Kotak, PNB, BoB, …)

Routes:
  POST /extract           — standard full-pipeline extraction
  POST /extract-enhanced  — YOLO-crop + region-wise OCR (images only)
  POST /export-xlsx       — export extracted JSON to Excel
"""

from __future__ import annotations

import io
import time
import logging
from typing import Optional

import cv2
import numpy as np

from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse

from app.models.schemas import (
    BankStatementResponse,
    OCRMetadata,
    ExtractionStatus,
)
from app.services.ocr_service import get_ocr_service, OCRService
from app.services.bank_extractor import BankStatementExtractor

logger = logging.getLogger(__name__)
router = APIRouter()

_ocr: Optional[OCRService] = None


def get_ocr() -> OCRService:
    global _ocr
    if _ocr is None:
        _ocr = get_ocr_service()
    return _ocr


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _is_pdf(content_type: str, contents: bytes) -> bool:
    """Detect PDF by content-type or magic bytes."""
    if "pdf" in content_type.lower():
        return True
    return contents[:4] == b"%PDF"


def _split_pdf_to_images(pdf_bytes: bytes, dpi: int = 200) -> list[bytes]:
    """
    Convert each page of a PDF to a PNG image byte array.
    Requires: pip install pdf2image  (which needs poppler on PATH)
    Falls back gracefully if pdf2image is not installed.
    """
    try:
        from pdf2image import convert_from_bytes
        pil_images = convert_from_bytes(pdf_bytes, dpi=dpi, fmt="png")
        result = []
        for img in pil_images:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result.append(buf.getvalue())
        return result
    except ImportError:
        logger.warning("pdf2image not installed — falling back to single-page PDF OCR")
        return []
    except Exception as e:
        logger.error(f"PDF→image conversion failed: {e}")
        return []


def _ocr_page(
    ocr: OCRService,
    page_bytes: bytes,
    content_type: str = "image/png",
) -> tuple[str, float, dict]:
    """Run OCR on a single page and return (text, confidence, meta)."""
    raw_text, confidence, ocr_meta = ocr.extract_from_file(page_bytes, content_type)
    return raw_text or "", confidence or 0.0, ocr_meta or {}


def _build_response(
    data,
    confidence: float,
    ocr_meta: dict,
    elapsed: float,
    engine_label: str = "ensemble",
    pages: int = 1,
) -> BankStatementResponse:
    status = (
        ExtractionStatus.SUCCESS
        if (data.transaction_count or 0) > 0
        else ExtractionStatus.PARTIAL
    )
    # pages_processed may not exist in older schemas — fall back to caller-supplied value
    pages_actual = getattr(data, "pages_processed", None) or pages
    return BankStatementResponse(
        status=status,
        message=(
            f"Extracted {data.transaction_count or 0} transactions "
            f"from {data.bank_name or 'unknown bank'} "
            f"({pages_actual} page{'s' if pages_actual != 1 else ''})"
        ),
        confidence_score=round(confidence, 3),
        extracted_data=data,
        ocr_metadata=OCRMetadata(
            engine_used=engine_label,
            confidence=round(confidence, 3),
            paddle_regions=ocr_meta.get("paddle_regions", 0),
            easy_regions=ocr_meta.get("easy_regions", 0),
            merged_regions=ocr_meta.get("merged_regions", 0),
            processing_time_ms=round(elapsed * 1000),
            pages_processed=pages_actual,
        ),
        processing_time_seconds=round(elapsed, 3),
    )


# ─────────────────────────────────────────────────────────────────────────────
# /extract  — Universal pipeline
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/extract", response_model=BankStatementResponse)
async def extract_bank_statement(file: UploadFile = File(...)):
    """
    Universal bank statement extractor.
    Accepts PDF (single or multi-page) or image (JPG/PNG/TIFF/BMP/WEBP).
    Returns fully structured transaction data for any bank.
    """
    start = time.time()
    contents = await file.read()
    content_type = file.content_type or "application/octet-stream"

    ocr = get_ocr()
    extractor = BankStatementExtractor()

    try:
        if _is_pdf(content_type, contents):
            # ── Multi-page PDF path ─────────────────────────────────────
            page_images = _split_pdf_to_images(contents)

            if page_images:
                # Best path: convert each page to image → OCR → extract per page
                all_text_parts: list[str] = []
                conf_scores: list[float] = []
                last_meta: dict = {}
                data = None  # accumulates BankStatementData across pages

                for page_idx, page_bytes in enumerate(page_images):
                    logger.info(f"Processing PDF page {page_idx + 1}/{len(page_images)}")
                    raw_text, conf, ocr_meta = _ocr_page(ocr, page_bytes, "image/png")

                    if raw_text:
                        all_text_parts.append(raw_text)
                        conf_scores.append(conf)
                        last_meta = ocr_meta

                        raw_boxes = ocr_meta.get("raw_boxes", [])
                        if raw_boxes:
                            data = extractor.extract_from_boxes(
                                raw_boxes,
                                full_text=raw_text,
                                page_count=page_idx + 1,
                                existing_data=data,
                            )
                        else:
                            # Text path for this page
                            if data is None:
                                data = extractor.extract(raw_text, page_count=1)
                                # extract() already parsed txns — don't re-add
                            else:
                                # Always run header extraction on every page:
                                # ICICI MICR on pg2, BoB IFSC/MICR on pg4 etc.
                                extractor._extract_header_fields(raw_text, data)
                                page_txs = extractor._parse_transactions_text(raw_text)
                                if data.transactions is None:
                                    data.transactions = []
                                data.transactions.extend(page_txs)
                                pages_so_far = getattr(data, "pages_processed", 0) or 0
                                try:
                                    data.pages_processed = pages_so_far + 1
                                except Exception:
                                    pass

                if data is None:
                    raise HTTPException(400, "OCR produced no text from PDF pages")

                full_text = "\n".join(all_text_parts)
                extractor.finalize(data, full_text)
                avg_conf = sum(conf_scores) / len(conf_scores) if conf_scores else 0.5
                pages_done = len(page_images)

            else:
                # pdf2image not available → send raw PDF bytes to OCR service
                raw_text, conf, ocr_meta = ocr.extract_from_file(contents, content_type)
                if not raw_text:
                    raise HTTPException(400, "OCR failed to extract text from PDF")

                raw_boxes = ocr_meta.get("raw_boxes", [])
                if raw_boxes:
                    data = extractor.extract_from_boxes(
                        raw_boxes, full_text=raw_text, page_count=1
                    )
                    extractor.finalize(data, raw_text)
                else:
                    data = extractor.extract(raw_text, page_count=ocr_meta.get("pages_processed", 1))

                avg_conf   = conf
                last_meta  = ocr_meta
                pages_done = ocr_meta.get("pages_processed", 1)

        else:
            # ── Single image path ───────────────────────────────────────
            raw_text, conf, ocr_meta = ocr.extract_from_file(contents, content_type)
            if not raw_text:
                raise HTTPException(400, "OCR failed to extract text from image")

            raw_boxes = ocr_meta.get("raw_boxes", [])
            if raw_boxes:
                data = extractor.extract_from_boxes(
                    raw_boxes, full_text=raw_text, page_count=1
                )
            else:
                data = extractor.extract(raw_text, page_count=1)

            extractor.finalize(data, raw_text)
            avg_conf   = conf
            last_meta  = ocr_meta
            pages_done = 1

        return _build_response(
            data, avg_conf, last_meta,
            elapsed=time.time() - start,
            engine_label="universal-v2",
            pages=pages_done,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bank statement extraction failed: {e}", exc_info=True)
        raise HTTPException(500, str(e))


# ─────────────────────────────────────────────────────────────────────────────
# /extract-enhanced  — YOLO-crop + region-wise OCR  (images only)
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/extract-enhanced", response_model=BankStatementResponse)
async def extract_bank_enhanced(file: UploadFile = File(...)):
    """
    Enhanced image pipeline:
      1. Detect header / table / footer regions via heuristic crop
      2. Preprocess each crop (grayscale, denoise, upscale)
      3. Run OCR per region for cleaner text
      4. Extract header fields from header crop, transactions from table crop
      For PDFs, falls back to the standard multi-page pipeline.
    """
    start = time.time()
    contents = await file.read()
    content_type = file.content_type or "application/octet-stream"

    ocr = get_ocr()
    extractor = BankStatementExtractor()

    try:
        # ── Try to decode as image ──────────────────────────────────────
        nparr = np.frombuffer(contents, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if img is not None:
            # Image path — region-crop pipeline
            h, w = img.shape[:2]
            crops = {
                "header": img[0:int(h * 0.20),          0:w],
                "table":  img[int(h * 0.20):int(h * 0.88), 0:w],
                "footer": img[int(h * 0.88):h,           0:w],
            }

            text_by_region: dict[str, str] = {}
            conf_scores: list[float] = []
            table_boxes: list = []
            last_meta: dict = {}

            for region_name, crop in crops.items():
                if crop.size == 0:
                    continue
                gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                scale = max(1.0, 800 / max(gray.shape[0], 1))
                if scale > 1.0:
                    gray = cv2.resize(gray, None, fx=scale, fy=scale,
                                      interpolation=cv2.INTER_CUBIC)
                gray = cv2.fastNlMeansDenoising(gray, h=10)
                region_bgr = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
                _, buf = cv2.imencode(".png", region_bgr)

                raw, conf, meta = ocr.extract_from_file(buf.tobytes(), "image/png")
                text_by_region[region_name] = raw or ""
                if region_name == "table":
                    table_boxes = meta.get("raw_boxes", [])
                    last_meta = meta
                if conf:
                    conf_scores.append(conf)

            full_text = "\n".join([
                text_by_region.get("header", ""),
                text_by_region.get("table",  ""),
                text_by_region.get("footer", ""),
            ])
            avg_conf = sum(conf_scores) / len(conf_scores) if conf_scores else 0.5

            if not full_text.strip():
                raise HTTPException(400, "OCR returned no text — try a higher-resolution image")

            # Extract using boxes if available, else text
            if table_boxes:
                data = extractor.extract_from_boxes(
                    table_boxes, full_text=full_text, page_count=1
                )
            else:
                data = extractor.extract(full_text, page_count=1)

            extractor.finalize(data, full_text)

            # ── Save annotated debug image ──────────────────────────────
            try:
                import os
                from pathlib import Path
                out_dir = Path("yolo_outputs")
                out_dir.mkdir(exist_ok=True)
                ts = int(time.time() * 1000)
                annotated = img.copy()
                COLOR_MAP = {
                    "header": (255, 100, 30),
                    "table":  (180, 30, 255),
                    "footer": (30, 200, 80),
                }
                for region_name, (y1, y2) in [
                    ("header", (0, int(h * 0.20))),
                    ("table",  (int(h * 0.20), int(h * 0.88))),
                    ("footer", (int(h * 0.88), h)),
                ]:
                    color = COLOR_MAP[region_name]
                    cv2.rectangle(annotated, (0, y1), (w, y2), color, 3)
                    cv2.putText(annotated, region_name, (10, y1 + 30),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)
                cv2.imwrite(str(out_dir / f"bank_{ts}.jpg"), annotated)
            except Exception:
                pass  # debug output failure should not break the response

            return _build_response(
                data, avg_conf, last_meta,
                elapsed=time.time() - start,
                engine_label="yolo-crop+ensemble",
                pages=1,
            )

        else:
            # PDF or non-decodable → delegate to standard multi-page pipeline
            # Re-use the /extract logic directly
            logger.info("Enhanced endpoint: non-image input, delegating to standard pipeline")
            file.seek(0) if hasattr(file, "seek") else None  # reset if possible
            # Re-build a fake UploadFile-like call using contents directly
            ocr_out = ocr.extract_from_file(contents, content_type)
            raw_text, conf, ocr_meta = ocr_out

            if not raw_text:
                raise HTTPException(400, "OCR failed to extract text")

            raw_boxes = ocr_meta.get("raw_boxes", [])
            if raw_boxes:
                data = extractor.extract_from_boxes(raw_boxes, full_text=raw_text)
            else:
                data = extractor.extract(raw_text)
            extractor.finalize(data, raw_text)

            return _build_response(
                data, conf, ocr_meta,
                elapsed=time.time() - start,
                engine_label="pdf-fallback",
                pages=ocr_meta.get("pages_processed", 1),
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Enhanced extraction failed: {e}", exc_info=True)
        raise HTTPException(500, str(e))


# ─────────────────────────────────────────────────────────────────────────────
# /export-xlsx
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/export-xlsx")
async def export_bank_xlsx(data: dict):
    """
    Export extracted bank statement data to a formatted Excel file.
    Accepts the extracted_data JSON object from /extract or /extract-enhanced.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Bank Statement"

    TITLE_FONT   = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    TITLE_FILL   = PatternFill("solid", fgColor="1B4332")
    SECTION_FONT = Font(bold=True, color="1B4332", size=10, name="Calibri")
    SECTION_FILL = PatternFill("solid", fgColor="D1FAE5")
    LABEL_FONT   = Font(bold=True, size=9, color="374151", name="Calibri")
    VALUE_FONT   = Font(size=10, name="Calibri")
    HDR_FONT     = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    HDR_FILL     = PatternFill("solid", fgColor="065F46")
    thin         = Side(style="thin", color="D1D5DB")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
    LEFT         = Alignment(horizontal="left", vertical="center", wrap_text=True)
    CENTER       = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    r = 1

    def title_row(text: str):
        nonlocal r
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font = TITLE_FONT; c.fill = TITLE_FILL
        c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[r].height = 24
        r += 1

    def section_row(text: str):
        nonlocal r
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font = SECTION_FONT; c.fill = SECTION_FILL
        c.alignment = LEFT; c.border = BORDER
        r += 1

    def kv(key: str, val):
        nonlocal r
        kc = ws.cell(row=r, column=1, value=key)
        kc.font = LABEL_FONT; kc.alignment = LEFT; kc.border = BORDER
        vc = ws.cell(row=r, column=2, value=val if val is not None else "—")
        vc.font = VALUE_FONT; vc.alignment = LEFT; vc.border = BORDER
        r += 1

    title_row(f"Bank Statement — {data.get('bank_name') or 'Unknown Bank'}")

    section_row("Account Details")
    kv("Bank Name",       data.get("bank_name"))
    kv("Account Holder",  data.get("account_holder"))
    kv("Account Number",  data.get("account_number"))
    kv("Account Type",    data.get("account_type"))
    kv("IFSC Code",       data.get("ifsc_code"))
    kv("MICR Code",       data.get("micr_code"))
    kv("Period From",     data.get("statement_period_from"))
    kv("Period To",       data.get("statement_period_to"))
    kv("Pages Processed", data.get("pages_processed"))

    section_row("Balances & Analytics")
    kv("Opening Balance",   data.get("opening_balance"))
    kv("Closing Balance",   data.get("closing_balance"))
    kv("Total Debits",      data.get("total_debits"))
    kv("Total Credits",     data.get("total_credits"))
    kv("Largest Debit",     data.get("largest_debit"))
    kv("Largest Credit",    data.get("largest_credit"))
    kv("Transaction Count", data.get("transaction_count"))

    transactions = data.get("transactions") or []
    if transactions:
        r += 1
        section_row(f"Transactions ({len(transactions)})")
        headers = ["Date", "Description", "Ref No", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CENTER; c.border = BORDER
        r += 1

        RED_FONT   = Font(size=9, color="DC2626", name="Calibri")
        GREEN_FONT = Font(size=9, color="059669", name="Calibri")
        MONO_FONT  = Font(size=9, name="Consolas")

        for tx in transactions:
            vals = [
                tx.get("date"),
                tx.get("description"),
                tx.get("ref_no"),
                tx.get("debit"),
                tx.get("credit"),
                tx.get("balance"),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=val if val is not None else "")
                c.alignment = LEFT; c.border = BORDER
                if ci == 4 and val:  c.font = RED_FONT
                elif ci == 5 and val: c.font = GREEN_FONT
                elif ci == 6:        c.font = MONO_FONT
                else:                c.font = Font(size=9, name="Calibri")
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    bank = (data.get("bank_name") or "bank").replace(" ", "_").lower()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={bank}_statement.xlsx"},
    )